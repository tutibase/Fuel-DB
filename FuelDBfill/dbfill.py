from faker import Faker
import random
import mysql.connector
import openpyxl
from mysql.connector import Error
from config import password


def export_to_sql():
    fake = Faker('ru_RU')
    try:
        connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password=password,
            database="fueldb",
        )

        rows_count = 0
        cursor = connection.cursor()

        query = "DELETE FROM `Driver-Fuel tanker`"
        cursor.execute(query)
        connection.commit()
        print(f"Все строки таблицы `Driver-Fuel tanker` были удалены.")

        query = "DELETE FROM `Fuel tanker-Fuel`"
        cursor.execute(query)
        connection.commit()
        print(f"Все строки таблицы `Fuel tanker-Fuel` были удалены.")

        query = "DELETE FROM `Technical Specialist-Fuel tanker`"
        cursor.execute(query)
        connection.commit()
        print(f"Все строки таблицы `Technical Specialist-Fuel tanker` были удалены.")

        query = "DELETE FROM `Fuel container`"
        cursor.execute(query)
        connection.commit()
        print(f"Все строки таблицы `Fuel container` были удалены.")

        query = "DELETE FROM `Fuel tanker-Gas station`"
        cursor.execute(query)
        connection.commit()
        print(f"Все строки таблицы `Fuel tanker-Gas station` были удалены.")

        query = "DELETE FROM `Technical Specialist-Fuel containers`"
        cursor.execute(query)
        connection.commit()
        print(f"Все строки таблицы `Technical Specialist-Fuel containers` были удалены.")

        # Таблиц уровня 1
        # заполнение таблицы Driver
        with connection.cursor() as cursor:
            for i in range(1, 501):
                cursor.execute("REPLACE INTO Driver (driver_id, Surname, Name, Middle_name,"
                               "Phone_number) VALUES (%s, %s, %s, %s, %s)",
                               (i, fake.last_name_male(), fake.first_name_male(), fake.middle_name_male(),
                                fake.phone_number()))
                connection.commit()
                rows_count += 1

        # заполнение таблицы Fuel tanker
        with connection.cursor() as cursor:
            file_to_read = openpyxl.load_workbook('data.xlsx', data_only=True)
            sheet = file_to_read['Лист2']
            for row in range(1, sheet.max_row + 1):
                value = sheet.cell(row, 1).value
                cursor.execute("REPLACE INTO `Fuel tanker` (Fuel_tanker_id, Model) VALUES (%s, %s)",
                               (row, value))
                connection.commit()
                rows_count += 1

        # заполнение таблицы Gas station
        with connection.cursor() as cursor:
            for i in range(1, 2001):
                cursor.execute("REPLACE INTO `Gas station` (Gas_station_id, Address) VALUES (%s, %s)",
                               (i, fake.address()))
                connection.commit()
                rows_count += 1

        # заполнение Fuel
        fuel_types = ['АИ-80', 'АИ-92', 'АИ-95', 'АИ-98', 'АИ-100', 'АИ-101',
                      'АИ-102', 'Дизельное топливо', 'Биодизельное топливо', 'Керосин']
        with connection.cursor() as cursor:
            cursor.executemany("REPLACE INTO Fuel (Fuel_id, Type) VALUES (%s, %s)",
                               [(id_num, fuel_type) for id_num, fuel_type in enumerate(fuel_types, start=1)])
            connection.commit()
        rows_count += 10

        # заполнение таблицы Technical Specialist
        with connection.cursor() as cursor:
            for i in range(1, 1001):
                cursor.execute("REPLACE INTO `Technical Specialist` (Technical_Specialist_id, "
                               "Surname, Name, Middle_name, Phone_number) VALUES (%s, %s, %s, %s, %s)",
                               (i, fake.last_name_male(), fake.first_name_male(), fake.middle_name_male(),
                                fake.phone_number()))
                connection.commit()
                rows_count += 1

        # Таблиц уровня 2
        # заполняем Driver-Fuel tanker
        with connection.cursor() as cursor:
            row_id = 1
            for Fuel_tanker_id in range(1, 541):  # для каждого бензовоза
                driver_count = random.randint(10, 20)  # получаем число водителей
                driver_ids = list(range(1, 501))  # заполняем список айдишников водителей
                for i in range(driver_count):  # цикл по количеству водителей
                    driver_id = driver_ids[random.randint(0, len(driver_ids)-1)]  # получаем id водителя как рандомное число из списка
                    driver_ids.remove(driver_id)  # удаляем это число, чтобы не было повторов
                    cursor.execute("REPLACE INTO `Driver-Fuel tanker` (`Driver-Fuel_tanker_id`, Driver_id, "
                                   "Fuel_tanker_id) VALUES (%s, %s, %s)", (row_id, driver_id, Fuel_tanker_id))
                    row_id += 1
                    rows_count += 1
            connection.commit()

        # заполняем Fuel tanker-Fuel
        with connection.cursor() as cursor:
            row_id = 1
            for Fuel_tanker_id in range(1, 541):
                fuel_count = random.randint(2, 7)
                fuel_ids = list(range(1, 11))
                for i in range(fuel_count):
                    fuel_id = fuel_ids[random.randint(0, len(fuel_ids) - 1)]
                    fuel_ids.remove(fuel_id)
                    cursor.execute("REPLACE INTO `Fuel tanker-Fuel` (`Fuel_tanker-Fuel_id`, Fuel_id, "
                                   "Fuel_tanker_id, Capacity) VALUES (%s, %s, %s, %s)",
                                   (row_id, fuel_id, Fuel_tanker_id, random.uniform(1200, 8000)))
                    row_id += 1
                    rows_count += 1
            connection.commit()

        # заполняем Technical Specialist-Fuel tanker
        with connection.cursor() as cursor:
            row_id = 1
            for Technical_Specialist_id in range(1, 1001):
                fuel_tanker_count = random.randint(1, 12)
                fuel_tanker_ids = list(range(1, 541))
                for i in range(fuel_tanker_count):
                    fuel_tanker_id = fuel_tanker_ids[random.randint(0, len(fuel_tanker_ids) - 1)]
                    fuel_tanker_ids.remove(fuel_tanker_id)
                    cursor.execute("REPLACE INTO `Technical Specialist-Fuel tanker` (`Technical_Specialist-Fuel_tanker_id`,"
                                   " Technical_Specialist_id, Fuel_tanker_id) VALUES (%s, %s, %s)",
                                   (row_id, Technical_Specialist_id, fuel_tanker_id))
                    row_id += 1
                    rows_count += 1
            connection.commit()

        # заполняем Fuel container
        with connection.cursor() as cursor:
            row_id = 1
            for Gas_station_id in range(1, 2001):
                fuel_containers_count = random.randint(20, 30)
                for i in range(fuel_containers_count):
                    capacity = random.uniform(20000, 40000)
                    fuel_amount = random.uniform(0, capacity)
                    cursor.execute(
                        "REPLACE INTO `Fuel container` (`Fuel_container_id`, Capacity, "
                        "Fuel_amount, Gas_station_id, Fuel_id) VALUES (%s, %s, %s, %s, %s)",
                        (row_id, capacity, fuel_amount, Gas_station_id, random.randint(1, 10)))
                    row_id += 1
                    rows_count += 1
            connection.commit()

        # Таблицы уровня 3
        # заполняем Fuel tanker-Gas station
        with connection.cursor() as cursor:
            row_id = 1

            cursor.execute('SELECT `Fuel_tanker-Fuel_id` FROM `Fuel tanker-Fuel`')
            delivery_values_list = [item[0] for item in cursor.fetchall()]

            for Gas_station_id in range(1, 2001):
                delivery_count = random.randint(50, 100)
                delivery_ids = delivery_values_list.copy()
                for i in range(delivery_count):
                    delivery_id = delivery_ids[random.randint(0, len(delivery_ids) - 1)]
                    delivery_ids.remove(delivery_id)
                    cursor.execute("REPLACE INTO `Fuel tanker-Gas station` (`Fuel_tanker-Gas_station_id`, "
                                   "`Fuel_tanker-Fuel_id`, Gas_station_id, Visit_date, Quantity) VALUES (%s, %s, %s, %s, %s)",
                                   (row_id, delivery_id, Gas_station_id, fake.date_between('-2y', 'today'), random.uniform(400, 3000)))
                    row_id += 1
                    rows_count += 1
            connection.commit()

        # заполняем Technical Specialist-Fuel containers
        with connection.cursor() as cursor:
            row_id = 1

            cursor.execute('SELECT `Fuel_tanker-Fuel_id` FROM `Fuel tanker-Fuel`')
            delivery_values_list = [item[0] for item in cursor.fetchall()]

            for Technical_Specialist_id in range(1, 1001):
                fuel_containers_count = random.randint(1, 10)
                fuel_containers_ids = delivery_values_list.copy()
                for i in range(fuel_containers_count):
                    fuel_containers_id = fuel_containers_ids[random.randint(0, len(fuel_containers_ids) - 1)]
                    fuel_containers_ids.remove(fuel_containers_id)
                    cursor.execute(
                        "REPLACE INTO `Technical Specialist-Fuel containers` (`Technical_Specialist-Fuel_containers_id`,"
                        " Technical_Specialist_id, Fuel_containers_id) VALUES (%s, %s, %s)",
                        (row_id, Technical_Specialist_id, fuel_containers_id))
                    row_id += 1
                    rows_count += 1
            connection.commit()

        #print table
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM `Technical Specialist-Fuel containers`")
        result = cursor.fetchall()
        for row in result:
            print(row)

        print("Количество записей в БД: " + str(rows_count))
    except Error as e:
        print(e)


export_to_sql()
